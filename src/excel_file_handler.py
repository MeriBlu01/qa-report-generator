import openpyxl
import logging


class ExcelFileHandler:


    def __init__(self, file_path, workbook_name):
        self.file_path = file_path
        self.workbook = None
        self.workbook_name = workbook_name
        self.sheet_count = 0


    def load_workbook(self):
        try:
            self.workbook = openpyxl.load_workbook(filename=self.file_path, data_only=True)
            self.sheet_count = self.get_sheet_count()  # Get and set the sheet count
        except Exception as e:
            logging.error(f"Error loading workbook: {self.workbook_name}. Please ensure the file is in the same location as the executable.")
            raise e
    

    def get_sheet_count(self):
        """Extract the number of sheets that a workbook has."""
        return len(self.workbook.sheetnames)
    

    def get_sheet_name(self, sheet_idx):
        """Get the name of a specific sheet by index."""
        if not self.workbook:
            raise ValueError("Workbook is not loaded. Please load the workbook first.")
        return self.workbook.sheetnames[sheet_idx]
    
    
    def get_sheet_by_index(self, sheet_idx):
        """Get a specific sheet by index."""
        if not self.workbook:
            raise ValueError("Workbook is not loaded. Please load the workbook first.")
        return self.workbook.worksheets[sheet_idx]
    

    def get_sheets_by_range(self, num_sheets=1):
        """Retrieves the first 'num_sheets' from the loaded configuration workbook.
        Returns them as a list of sheets.
        
        :param num_sheets: The number of sheets to retrieve (default is 1).
        :return: List of sheet objects.
        """
        if not self.workbook:
            raise ValueError("Workbook is not loaded. Please load the workbook first.")

        # Get the specified number of sheets from the workbook
        sheets = self.workbook.worksheets[:num_sheets]

        return sheets


    def get_sheet_by_name(self, sheet_name):
        """Retrieves a worksheet by its name from the loaded workbook.

        :param sheet_name: The name of the sheet to retrieve.
        :returns: The worksheet object corresponding to the specified sheet name.
        """
        if not self.workbook:
            self.load_workbook()
        try:
            return self.workbook[sheet_name]
        except KeyError:
            logging.error(f"Sheet '{sheet_name}' not found in {self.file_path}")
            raise

    
    def get_sheet_index_by_name(self, sheet_name):
        """Retrieves the worksheet index by its name from the loaded workbook.

        :param sheet_name: The name of the sheet to search.
        :returns: The index of the worksheet corresponding to the specified sheet name.
        """

        if not self.workbook:
            self.load_workbook()
        
        # Get the sheet names
        sheet_names = self.workbook.sheetnames
        
        # Find the index of the specified sheet name
        return sheet_names.index(sheet_name)
    

    @staticmethod
    def find_table_dimensions(sheet, initial_row=0):
        """Find the rows and columns range of the table in the sheet.

        :param sheet: The sheet where the table is located.
        :param initial_row: The initial row the table is located in the sheet (default is 1).
        :return: The table dimensions (rows, columns).
        """

        rows = max((cell.row for cell in sheet['A'] if cell.value is not None), default=0)
        cols = max((cell.column for cell in sheet[1+initial_row] if cell.value is not None), default=0)
        return rows, cols
    

    @staticmethod
    def extract_table_data(sheet, final_row, final_col, initial_row=2):
        """Extract table data from sheet into a list of lists.

        :param sheet: The sheet where the table is located.
        :param final_row: Is equal to the table rows dimension
        :param final_col: Is equal to the table columns dimension
        :param initial_row: The initial row default is 2 because it's excluding the table header row.
        :return: The table data.
        """
        return [
            [sheet.cell(row=r, column=c).value for r in range(initial_row, final_row + 1)]
            for c in range(1, final_col + 1)
        ]
    

    @staticmethod
    def lookup_function(table, search_value, search_col_idex):
        """Lookup function: finds the index of the value within a column\list.
        
        :returns: Returns the row index where the value was found.
        """
        if search_value in table[search_col_idex]:
            # Get the row index of the value in the target lookup column
            return_row_idx = table[search_col_idex].index(search_value)
            return return_row_idx
        else:
            return None


    def save_workbook(self, save_path=None):
        """Save the workook into a specified directory path.
        """
        if save_path is None:
            save_path = self.file_path
        self.workbook.save(save_path)
        logging.info(f"Workbook saved: {save_path}")