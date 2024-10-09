import openpyxl
import logging
from datetime import datetime

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

from .config_variables import data_header_table


class ReportGenerator:
    def __init__(self):
        self.report_filename = ''
        self.creation_flag = True
        self.reason = ''

    def generate_filename(self):
        """
        Generates a timestamped filename for the report.
        """
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
        return f"Report_{timestamp}.xlsx"
      

    def create_table_n_apply_table_style(self, table_range):

        # Apply table style
        style_name = "TableStyleMedium2"
        style = TableStyleInfo(
            name=style_name,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table = Table(displayName="ReportTable", ref=table_range)
        table.tableStyleInfo = style

        return table


    def english_date_style(self, ws, wb, data_table):
        """Ensure "english_date" style exists"""
        if "english_date" not in wb.named_styles:
            english_date = NamedStyle(name="english_date", number_format="MM-DD-YYYY")
            wb.add_named_style(english_date)

        # Apply date style to the date columns
        for row in ws.iter_rows(min_row=2, max_row=len(data_table), min_col=4, max_col=4):
            for cell in row:
                cell.style = "english_date"


    def generate_report(self, data, config):
        """Generate the report in a new workbook.
        
        :param data: List of dictionaries.
        :param config: Contains table settings.
        """
        
        # Create a new workbook for the final report and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        self.report_filename = self.generate_filename()
        data_table_length = config.get("data_table_length")

        if self.creation_flag:
            if data_table_length > 1:

                # Insert data into 'Report' sheet
                for i, row in enumerate(data, start=1):
                    for j, value in enumerate(row, start=1):
                        ws.cell(row=i, column=j, value=value)

                # Define table range
                # Given column number
                col_number = len(data_header_table)
                # Get the corresponding column letter
                column_letter = get_column_letter(col_number)
                table_range = f"A1:{column_letter}{data_table_length}"

                # Create the table and apply the style
                report_data_table = self.create_table_n_apply_table_style(table_range)

                # Ensure "english_date" style exists
                self.english_date_style(ws, wb, data)

                # Add the table to the worksheet
                ws.add_table(report_data_table)

                # Save the workbook to a file
                wb.save(self.report_filename)

                
                print(f"\nNew Report saved as: {self.report_filename}")
                logging.warning(f"New Report saved as: {self.report_filename}")
            elif data_table_length == None:
                logging.warning(f"There're no records.")
                self.creation_flag = False
            else: 
                logging.warning(f"There're no records that meet the given criteria: {self.reason}.")
                self.creation_flag = False
        else:
            self.creation_flag = False
            logging.warning(f"Report file wasn't created.")
            print(f"\nReport file wasn't created, please review the 'error_log.txt' file for more details.")
        